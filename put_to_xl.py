from re import S
import openpyxl as p
import os


def add_to_xl(st_list):
    f_name = 'stockxl.xlsx'
    s_list = sorted(st_list, key = lambda x: x[0])

    if os.path.isfile(f_name):
        wb = p.load_workbook(f_name)
        sheet = wb.active
    else:
        wb = p.Workbook()
        sheet = wb.active
        sheet.title = 'Stocks'
        sheet['A1'] = 'Name'
        sheet['B1'] = 'Price'
        sheet['C1'] = 'P/E'
        sheet['D1'] = 'EY %'
        sheet['E1'] = 'Market Cap'
        sheet['F1'] = 'Dividend Yield'
        wb.save(f_name)

    starting_row = sheet.max_row + 1

    for lst in s_list:
        starting_clmn = 1
        for item in lst:
            sheet.cell(row = starting_row, column = starting_clmn).value = item
            starting_clmn += 1
        starting_row += 1
    
    wb.save(f_name)
